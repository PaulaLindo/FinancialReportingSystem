"""
Export Center — Excel, CSV, archive, and session-backed PDF payloads.

All exports require CFO-finalized sessions in a locked reporting period.
"""

from __future__ import annotations

import csv
import io
import json
import os
import zipfile
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font

from services.statement_validation_service import (
    group_mapped_accounts_for_statements,
    mapped_lines_from_metadata,
)
from utils.pdf_availability import resolve_pdf_availability
from utils.session_workflow import effective_workflow_status


def _amount(line: Dict[str, Any]) -> float:
    for key in ("current_amount", "amount", "net_balance", "balance", "value", "actual", "budget"):
        if key in line and line[key] is not None:
            try:
                return float(line[key])
            except (TypeError, ValueError):
                continue
    return 0.0


def _line_label(line: Dict[str, Any]) -> str:
    for key in ("account_name", "name", "account_desc", "description", "line_item", "category"):
        val = line.get(key)
        if val:
            return str(val)
    code = line.get("account_code") or line.get("code")
    return str(code or "Line")


def _pdf_row(line: Dict[str, Any], *, default_amount: Optional[float] = None) -> Dict[str, Any]:
    return {
        "Line Item": _line_label(line),
        "Amount": float(default_amount if default_amount is not None else _amount(line)),
    }


def _accounts_to_pdf_rows(block: Any) -> List[Dict[str, Any]]:
    if isinstance(block, dict):
        accounts = block.get("accounts") or []
        total = block.get("total")
        rows = [_pdf_row(a) for a in accounts if isinstance(a, dict)]
        if not rows and total is not None:
            rows = [{"Line Item": block.get("label") or "Total", "Amount": float(total)}]
        return rows
    if isinstance(block, list):
        return [_pdf_row(a) for a in block if isinstance(a, dict)]
    return []


def _normalize_budget_export_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Map budget session row keys to stable export column names."""
    budget = row.get("budget_amount")
    if budget is None:
        budget = row.get("budget")
    actual = row.get("actual_amount")
    if actual is None:
        actual = row.get("actual")
    variance_pct = row.get("variance_percentage")
    if variance_pct is None:
        variance_pct = row.get("variance_pct")
    return {
        "account_code": row.get("account_code") or "",
        "account_name": (
            row.get("account_description")
            or row.get("account_name")
            or row.get("expense_category")
            or ""
        ),
        "department": row.get("department") or "",
        "budget": budget,
        "actual": actual,
        "variance": row.get("variance"),
        "variance_pct": variance_pct,
        "grap_code": row.get("mapped_to_grap") or row.get("grap_code") or "",
        "variance_explanation": row.get("variance_explanation") or "",
    }


def _budget_export_rows(summary: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows = summary.get("budget_rows") or []
    return [_normalize_budget_export_row(r) for r in rows if isinstance(r, dict)]


def _compute_summary_ratios(
    *,
    total_assets: float,
    total_liabilities: float,
    net_assets: float,
    total_revenue: float,
    surplus: float,
) -> Dict[str, float]:
    """Key ratios expected by PDFService (matches legacy upload summary shape)."""
    total_equity = net_assets if net_assets else max(total_assets - total_liabilities, 0.0)
    return {
        "current_ratio": total_assets / total_liabilities if total_liabilities else 0.0,
        "debt_to_equity": total_liabilities / total_equity if total_equity else 0.0,
        "return_on_assets": surplus / max(total_assets, 1.0),
        "operating_margin": (surplus / total_revenue * 100.0) if total_revenue else 0.0,
    }


def _build_budget_pdf_results_from_summary(summary: Dict[str, Any]) -> Dict[str, Any]:
    """Map GRAP 24 budget session → PDF performance-focused layout."""
    rows = summary.get("budget_rows") or []
    expense_lines: List[Dict[str, Any]] = []
    for row in rows:
        if row.get("is_total_row") or row.get("is_subtotal_row"):
            continue
        desc = (row.get("account_description") or row.get("account_code") or "Line").strip()
        dept = (row.get("department") or "").strip()
        if dept and dept.lower() not in desc.lower():
            desc = f"{dept} — {desc}"
        budget_amt = float(row.get("budget_amount") or 0)
        actual_amt = float(row.get("actual_amount") or 0)
        expense_lines.append(
            {
                "Line Item": f"{desc} (Budget R{budget_amt:,.2f})",
                "Amount": actual_amt,
            }
        )

    total_budget = float(summary.get("total_budget") or 0)
    total_actual = float(summary.get("total_actual") or 0)
    total_variance = float(summary.get("total_variance") or (total_actual - total_budget))
    surplus_f = total_budget - total_actual

    ratios = _compute_summary_ratios(
        total_assets=total_actual,
        total_liabilities=0.0,
        net_assets=total_variance,
        total_revenue=total_budget,
        surplus=surplus_f,
    )

    return {
        "summary": {
            "total_assets": 0.0,
            "total_liabilities": 0.0,
            "net_assets": total_variance,
            "total_revenue": total_budget,
            "total_expenses": total_actual,
            "total_budget": total_budget,
            "total_actual": total_actual,
            "total_variance": total_variance,
            "surplus_deficit": surplus_f,
            "ratios": ratios,
            "session_id": summary.get("session_id"),
            "document_type": "budget_report",
            "filename": summary.get("filename"),
        },
        "sofp": {
            "assets": [{"Line Item": "Not applicable — budget report", "Amount": 0.0}],
            "liabilities": [{"Line Item": "Not applicable — budget report", "Amount": 0.0}],
            "net_assets": [{"Line Item": "Budget variance", "Amount": total_variance}],
        },
        "sofe": {
            "revenue": [{"Line Item": "Total approved budget", "Amount": total_budget}],
            "expenses": expense_lines
            or [{"Line Item": "Total actual expenditure", "Amount": total_actual}],
            "surplus": surplus_f,
        },
        "processed_at": datetime.now().isoformat(),
    }


def build_pdf_results_from_summary(summary: Dict[str, Any]) -> Dict[str, Any]:
    """Convert universal session summary → PDFService results dict."""
    doc_type = (summary.get("document_type") or "").strip().lower()
    if doc_type == "budget_report" and summary.get("budget_rows"):
        return _build_budget_pdf_results_from_summary(summary)

    fs = summary.get("financial_statements") or {}
    if not fs:
        mapped = mapped_lines_from_metadata(summary.get("metadata") or {})
        if mapped:
            fs = group_mapped_accounts_for_statements(mapped)

    sfp = fs.get("statement_of_financial_position") or {}
    perf = fs.get("statement_of_financial_performance") or {}

    assets = _accounts_to_pdf_rows(sfp.get("assets"))
    liabilities = _accounts_to_pdf_rows(sfp.get("liabilities"))
    equity_total = float((sfp.get("equity") or {}).get("total") or 0)
    liab_total = float((sfp.get("liabilities") or {}).get("total") or 0)
    assets_total = float((sfp.get("assets") or {}).get("total") or 0)
    net_assets_amount = assets_total - liab_total if assets_total or liab_total else equity_total

    revenue_rows = _accounts_to_pdf_rows(perf.get("revenue"))
    expense_rows = _accounts_to_pdf_rows(perf.get("expenses"))
    surplus = perf.get("surplus")
    if surplus is None:
        rev_t = float((perf.get("revenue") or {}).get("total") or 0)
        exp_t = float((perf.get("expenses") or {}).get("total") or 0)
        surplus = rev_t - exp_t

    total_revenue = float((perf.get("revenue") or {}).get("total") or 0)
    surplus_f = float(surplus or 0)
    existing_ratios = summary.get("ratios") or (summary.get("summary") or {}).get("ratios") or {}
    ratios = {
        **_compute_summary_ratios(
            total_assets=assets_total,
            total_liabilities=liab_total,
            net_assets=net_assets_amount,
            total_revenue=total_revenue,
            surplus=surplus_f,
        ),
        **(existing_ratios if isinstance(existing_ratios, dict) else {}),
    }

    return {
        "summary": {
            "total_assets": assets_total,
            "total_liabilities": liab_total,
            "net_assets": net_assets_amount,
            "total_revenue": total_revenue,
            "total_expenses": float((perf.get("expenses") or {}).get("total") or 0),
            "surplus_deficit": surplus_f,
            "ratios": ratios,
            "session_id": summary.get("session_id"),
            "document_type": summary.get("document_type"),
            "filename": summary.get("filename"),
        },
        "sofp": {
            "assets": assets or [{"Line Item": "Total assets", "Amount": assets_total}],
            "liabilities": liabilities or [{"Line Item": "Total liabilities", "Amount": liab_total}],
            "net_assets": [{"Line Item": "Net assets / equity", "Amount": net_assets_amount}],
        },
        "sofe": {
            "revenue": revenue_rows or [{"Line Item": "Total revenue", "Amount": 0}],
            "expenses": expense_rows or [{"Line Item": "Total expenses", "Amount": 0}],
            "surplus": float(surplus or 0),
        },
        "processed_at": datetime.now().isoformat(),
    }


class ExportCenterService:
    """Build downloadable exports from finalized universal sessions."""

    DOCUMENT_TYPES = ("balance_sheet", "income_statement", "budget_report")

    def __init__(self):
        from services.universal_workflow_service import UniversalWorkflowService

        self._workflow = UniversalWorkflowService()

    def _model(self, document_type: str):
        return self._workflow._get_model_for_document_type(document_type)

    def _service(self, document_type: str):
        from services.budget_report_service import BudgetReportService
        from services.flexible_balance_sheet_service import FlexibleBalanceSheetService
        from services.income_statement_service import IncomeStatementService

        services = {
            "balance_sheet": FlexibleBalanceSheetService,
            "income_statement": IncomeStatementService,
            "budget_report": BudgetReportService,
        }
        cls = services.get((document_type or "").strip().lower())
        return cls() if cls else None

    def load_session(self, session_id: str, document_type: str):
        model = self._model(document_type)
        return model.get_session(session_id) if model else None

    def session_is_exportable(self, session, document_type: str) -> Tuple[bool, str]:
        if not session:
            return False, "Session not found"
        status = effective_workflow_status(session)
        if status != "approved":
            return False, "Session must be CFO-approved before export"
        avail = resolve_pdf_availability(session_id=session.id, document_type=document_type)
        if not avail.get("period_locked"):
            return False, avail.get("reason") or "Reporting period is not locked"
        return True, ""

    def load_export_payload(self, session_id: str, document_type: str) -> Dict[str, Any]:
        ok, err = self.session_is_exportable(
            self.load_session(session_id, document_type),
            document_type,
        )
        if not ok:
            raise ValueError(err)

        service = self._service(document_type)
        if not service or not hasattr(service, "get_session_summary"):
            raise ValueError(f"No export service for document type: {document_type}")

        summary = service.get_session_summary(session_id)
        if not summary or summary.get("error"):
            raise ValueError(summary.get("error") or "Could not load session summary")

        summary.setdefault("session_id", session_id)
        summary.setdefault("document_type", document_type)
        return summary

    def list_exportable_sessions(self, *, limit: int = 50) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        seen: set = set()

        for doc_type in self.DOCUMENT_TYPES:
            model = self._model(doc_type)
            if not model or not getattr(model, "client", None):
                continue
            table = getattr(model, "table_name", None)
            if not table and doc_type == "balance_sheet":
                table = "balance_sheet_sessions"
            if not table:
                continue
            try:
                result = (
                    model.client.table(table)
                    .select("*")
                    .order("updated_at", desc=True)
                    .limit(max(limit * 4, 80))
                    .execute()
                )
            except Exception:
                continue

            for raw in result.data or []:
                try:
                    if doc_type == "balance_sheet":
                        from models.balance_sheet_models import _balance_sheet_session_from_row

                        sess = _balance_sheet_session_from_row(raw)
                    elif doc_type == "income_statement":
                        from models.income_statement_models import income_statement_session_from_row

                        sess = income_statement_session_from_row(raw)
                    else:
                        from models.budget_report_models import budget_report_session_from_row

                        sess = budget_report_session_from_row(raw)
                except Exception:
                    continue

                sid = getattr(sess, "id", None)
                if not sid or sid in seen:
                    continue
                if effective_workflow_status(sess) != "approved":
                    continue
                avail = resolve_pdf_availability(session_id=sid, document_type=doc_type)
                if not avail.get("period_locked"):
                    continue

                seen.add(sid)
                md = getattr(sess, "metadata", None) or {}
                rows.append(
                    {
                        "session_id": sid,
                        "document_type": doc_type,
                        "filename": getattr(sess, "filename", "") or "",
                        "period_name": avail.get("period_name") or md.get("period_name"),
                        "period_id": avail.get("period_id") or md.get("period_id"),
                        "updated_at": getattr(sess, "updated_at", None),
                        "approved_at": md.get("approved_at"),
                    }
                )

        rows.sort(key=lambda r: str(r.get("updated_at") or ""), reverse=True)
        return rows[:limit]

    def _sheet_title(self, base: str) -> str:
        return base[:31]

    def export_excel_bytes(self, summary: Dict[str, Any]) -> bytes:
        wb = Workbook()
        ws_info = wb.active
        ws_info.title = self._sheet_title("Session")
        bold = Font(bold=True)
        ws_info["A1"] = "Field"
        ws_info["B1"] = "Value"
        ws_info["A1"].font = bold
        ws_info["B1"].font = bold
        info_rows = [
            ("Session ID", summary.get("session_id")),
            ("Document type", summary.get("document_type")),
            ("Filename", summary.get("filename")),
            ("Status", (summary.get("metadata") or {}).get("workflow_status") or summary.get("status")),
            ("Exported at", datetime.now().isoformat()),
        ]
        for i, (k, v) in enumerate(info_rows, start=2):
            ws_info[f"A{i}"] = k
            ws_info[f"B{i}"] = v

        doc_type = (summary.get("document_type") or "").lower()
        if doc_type == "budget_report":
            self._write_table_sheet(
                wb,
                "Budget",
                _budget_export_rows(summary),
                [
                    "account_code",
                    "account_name",
                    "department",
                    "budget",
                    "actual",
                    "variance",
                    "variance_pct",
                    "grap_code",
                    "variance_explanation",
                ],
            )
        else:
            fs = summary.get("financial_statements") or {}
            if not fs:
                mapped = mapped_lines_from_metadata(summary.get("metadata") or {})
                fs = group_mapped_accounts_for_statements(mapped) if mapped else {}

            sfp = fs.get("statement_of_financial_position") or {}
            for section, title in (
                ("assets", "SFP Assets"),
                ("liabilities", "SFP Liabilities"),
                ("equity", "SFP Equity"),
            ):
                block = sfp.get(section) or {}
                accounts = block.get("accounts") if isinstance(block, dict) else block
                self._write_table_sheet(
                    wb,
                    title,
                    accounts or [],
                    ["account_code", "account_name", "grap_code", "amount", "debit", "credit"],
                )

            perf = fs.get("statement_of_financial_performance") or {}
            for section, title in (("revenue", "Revenue"), ("expenses", "Expenses")):
                block = perf.get(section) or {}
                accounts = block.get("accounts") if isinstance(block, dict) else block
                self._write_table_sheet(
                    wb,
                    title,
                    accounts or [],
                    ["account_code", "account_name", "grap_code", "amount"],
                )

        mapped = mapped_lines_from_metadata(summary.get("metadata") or {})
        self._write_table_sheet(
            wb,
            "Mappings",
            mapped,
            ["account_code", "account_name", "grap_code", "grap_category", "amount", "debit_balance", "credit_balance"],
        )

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _write_table_sheet(
        self,
        wb: Workbook,
        title: str,
        rows: List[Dict[str, Any]],
        columns: List[str],
    ) -> None:
        ws = wb.create_sheet(title=self._sheet_title(title))
        bold = Font(bold=True)
        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col)
            cell.font = bold
        for row_idx, row in enumerate(rows, start=2):
            if not isinstance(row, dict):
                continue
            for col_idx, col in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col, ""))

    def export_csv_bytes(self, summary: Dict[str, Any]) -> bytes:
        buf = io.StringIO()
        doc_type = (summary.get("document_type") or "").lower()
        if doc_type == "budget_report":
            rows = _budget_export_rows(summary)
            fieldnames = [
                "account_code",
                "account_name",
                "department",
                "budget",
                "actual",
                "variance",
                "variance_pct",
                "grap_code",
                "variance_explanation",
            ]
        else:
            rows = mapped_lines_from_metadata(summary.get("metadata") or {})
            fieldnames = [
                "account_code",
                "account_name",
                "grap_code",
                "grap_category",
                "amount",
                "debit_balance",
                "credit_balance",
                "debit",
                "credit",
            ]

        writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            if isinstance(row, dict):
                writer.writerow({k: row.get(k, "") for k in fieldnames})
        return buf.getvalue().encode("utf-8-sig")

    def export_archive_bytes(
        self,
        summary: Dict[str, Any],
        *,
        output_folder: Optional[str] = None,
        include_pdf_path: Optional[str] = None,
    ) -> bytes:
        sid = summary.get("session_id") or "session"
        doc_type = summary.get("document_type") or "document"
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = f"{doc_type}_{sid[:8]}_{stamp}"

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"{prefix}/session_summary.json", json.dumps(summary, indent=2, default=str))
            zf.writestr(f"{prefix}/export.csv", self.export_csv_bytes(summary))
            zf.writestr(f"{prefix}/export.xlsx", self.export_excel_bytes(summary))

            md = summary.get("metadata") or {}
            if md.get("variance_explanations"):
                zf.writestr(
                    f"{prefix}/variance_explanations.json",
                    json.dumps(md["variance_explanations"], indent=2),
                )

            if include_pdf_path and os.path.isfile(include_pdf_path):
                zf.write(include_pdf_path, arcname=f"{prefix}/financial_statements.pdf")
            elif output_folder:
                pdf_match = self.find_session_pdf(output_folder, sid, doc_type)
                if pdf_match:
                    zf.write(pdf_match, arcname=f"{prefix}/financial_statements.pdf")

        return buf.getvalue()

    def find_session_pdf(
        self,
        output_folder: str,
        session_id: str,
        document_type: Optional[str] = None,
    ) -> Optional[str]:
        if not output_folder or not os.path.isdir(output_folder):
            return None
        for name in os.listdir(output_folder):
            if not name.lower().endswith(".pdf"):
                continue
            meta_path = os.path.join(output_folder, f"{name}.meta.json")
            if not os.path.isfile(meta_path):
                continue
            try:
                with open(meta_path, encoding="utf-8") as fh:
                    meta = json.load(fh)
                if meta.get("session_id") == session_id:
                    if not document_type or meta.get("document_type") == document_type:
                        full = os.path.join(output_folder, name)
                        if os.path.isfile(full):
                            return full
            except (json.JSONDecodeError, OSError):
                continue
        return None


export_center_service = ExportCenterService()
